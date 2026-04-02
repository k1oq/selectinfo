import csv
import tempfile
import unittest
from io import StringIO
from pathlib import Path

from openpyxl import load_workbook

from _bootstrap import PROJECT_ROOT  # noqa: F401
from core.human_reports import (
    build_batch_summary_report,
    build_single_scan_report,
    default_report_path,
    write_batch_summary_report,
    write_single_scan_report,
)


class HumanReportsTests(unittest.TestCase):
    def _sample_single_result(self) -> dict:
        return {
            "target": "example.com",
            "scan_time": "2026-04-01T10:00:00",
            "scan_preset": "deep",
            "duration_seconds": 12.5,
            "tools_used": ["subfinder", "oneforall"],
            "wildcard": {"detected": False},
            "statistics": {"total_found": 5, "valid_count": 3, "filtered_count": 2},
            "tool_runs": {
                "subfinder": {
                    "status": "completed",
                    "return_code": 0,
                    "message": "ok",
                    "raw_count": 5,
                    "valid_count": 3,
                }
            },
            "subdomains": [{"subdomain": "www.example.com", "ip": ["1.1.1.1"], "alive_verified": True}],
            "port_scan": {
                "statistics": {"total_open_ports": 2},
                "hosts": {"1.1.1.1": [80, 443]},
            },
            "web_fingerprint": {
                "statistics": {"web_target_count": 1},
                "targets": [
                    {
                        "url": "https://www.example.com",
                        "subdomain": "www.example.com",
                        "port": 443,
                        "ip": "1.1.1.1",
                        "fingerprint_status": "identified",
                        "nmap": {"title": "Example", "server_header": "nginx", "service": "https"},
                    }
                ],
            },
            "directory_scan": {
                "statistics": {"interesting_path_count": 1},
                "targets": [
                    {
                        "url": "https://www.example.com",
                        "subdomain": "www.example.com",
                        "ip": "1.1.1.1",
                        "findings": [{"path": "/admin", "status": 200, "redirect": ""}],
                    }
                ],
            },
        }

    def _sample_batch_summary(self) -> dict:
        return {
            "scan_time": "2026-04-01T10:00:00",
            "scan_preset": "quick",
            "tools_used": ["subfinder"],
            "statistics": {
                "requested_domains": 2,
                "total_domains": 2,
                "success_count": 1,
                "error_count": 1,
                "no_valid_result_count": 0,
                "total_open_ports": 4,
                "total_web_targets": 2,
                "total_dirsearch_findings": 1,
            },
            "items": [
                {
                    "domain": "a.com",
                    "status": "success",
                    "valid_count": 2,
                    "total_found": 4,
                    "wildcard_detected": False,
                    "saved_path": "results/a.json",
                    "port_scan_status": "completed",
                    "open_port_count": 4,
                    "web_fingerprint_status": "completed",
                    "web_target_count": 2,
                    "directory_scan_status": "completed",
                    "dirsearch_finding_count": 1,
                    "message": "ok",
                },
                {
                    "domain": "b.com",
                    "status": "error",
                    "valid_count": 0,
                    "total_found": 0,
                    "wildcard_detected": False,
                    "saved_path": None,
                    "port_scan_status": "not_started",
                    "open_port_count": 0,
                    "web_fingerprint_status": "not_started",
                    "web_target_count": 0,
                    "directory_scan_status": "not_started",
                    "dirsearch_finding_count": 0,
                    "message": "boom",
                },
            ],
        }

    def test_single_scan_report_contains_key_sections(self):
        report = build_single_scan_report(
            self._sample_single_result(),
            source_path=Path(PROJECT_ROOT) / "results" / "example.json",
        )

        rows = list(csv.DictReader(StringIO(report)))
        self.assertEqual(rows[0]["分组"], "概览")
        self.assertEqual(rows[0]["名称"], "目标域名")
        self.assertEqual(rows[0]["数值"], "example.com")
        self.assertTrue(any(row["分组"] == "概览" and row["名称"] == "参数档位" and row["数值"] == "deep" for row in rows))
        self.assertTrue(any(row["分组"] == "工具状态" and row["名称"] == "subfinder" for row in rows))
        self.assertTrue(any(row["分组"] == "Web目标" and row["名称"] == "https://www.example.com" for row in rows))
        self.assertTrue(any(row["分组"] == "目录发现" and row["数值"] == "/admin" for row in rows))

    def test_batch_summary_report_contains_overview_and_failures(self):
        report = build_batch_summary_report(
            self._sample_batch_summary(),
            source_path=Path(PROJECT_ROOT) / "results" / "batch_summary.json",
        )

        rows = list(csv.DictReader(StringIO(report)))
        self.assertTrue(any(row["分组"] == "概览" and row["名称"] == "参数档位" and row["数值"] == "quick" for row in rows))
        self.assertTrue(any(row["分组"] == "统计" and row["名称"] == "成功数" and row["数值"] == "1" for row in rows))
        self.assertTrue(any(row["分组"] == "域名概览" and row["名称"] == "a.com" for row in rows))
        self.assertTrue(any(row["分组"] == "域名概览" and row["名称"] == "b.com" and row["状态"] == "error" for row in rows))

    def test_default_report_path_reuses_result_stem_with_xlsx_suffix(self):
        self.assertEqual(
            default_report_path(Path(PROJECT_ROOT) / "results" / "example.com_20260402_134217.json"),
            Path(PROJECT_ROOT) / "results" / "example.com_20260402_134217.xlsx",
        )

    def test_write_single_scan_report_defaults_to_multi_sheet_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_path = write_single_scan_report(
                self._sample_single_result(),
                source_path=Path(tmp) / "example.json",
            )

            self.assertEqual(output_path.suffix.lower(), ".xlsx")
            workbook = load_workbook(output_path)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["概览", "统计", "工具状态", "子域名", "端口扫描", "Web指纹", "目录扫描"],
                )
                self.assertEqual(workbook["概览"]["A2"].value, "目标域名")
                self.assertEqual(workbook["概览"]["B5"].value, "deep")
                self.assertEqual(workbook["子域名"]["A2"].value, "www.example.com")
                self.assertEqual(workbook["端口扫描"]["A2"].value, "1.1.1.1")
                self.assertEqual(workbook["目录扫描"]["B2"].value, "/admin")
            finally:
                workbook.close()

    def test_write_batch_summary_report_defaults_to_multi_sheet_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_path = write_batch_summary_report(
                self._sample_batch_summary(),
                source_path=Path(tmp) / "batch_summary.json",
            )

            self.assertEqual(output_path.suffix.lower(), ".xlsx")
            workbook = load_workbook(output_path)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["概览", "统计", "域名概览", "端口扫描", "Web指纹", "目录扫描"],
                )
                self.assertEqual(workbook["概览"]["B4"].value, "quick")
                self.assertEqual(workbook["域名概览"]["A2"].value, "a.com")
                self.assertEqual(workbook["Web指纹"]["B2"].value, "completed")
            finally:
                workbook.close()

    def test_write_single_scan_report_normalizes_explicit_csv_path_to_xlsx(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "example.summary.csv"
            written = write_single_scan_report(
                self._sample_single_result(),
                source_path=Path(tmp) / "example.json",
                output_path=output_path,
            )

            self.assertEqual(written, output_path.with_suffix(".xlsx"))
            workbook = load_workbook(written)
            try:
                self.assertEqual(workbook["概览"]["A2"].value, "目标域名")
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
