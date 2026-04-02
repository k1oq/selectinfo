"""
Human-friendly summary reports for scan results.
"""

from __future__ import annotations

import csv
import io
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from utils import load_json_file


CSV_HEADERS = ["序号", "分组", "名称", "状态", "数值", "详情1", "详情2", "详情3", "备注"]
WORKBOOK_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
WORKBOOK_HEADER_FONT = Font(bold=True)
WORKBOOK_CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


def default_report_path(source_path: Path | str) -> Path:
    path = Path(source_path)
    return path.with_suffix(".xlsx")


def write_csv_report(content: str, output_path: Path | str) -> Path:
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(
            "w",
            encoding="utf-8-sig",
            dir=str(destination.parent),
            prefix=f".{destination.name}.",
            suffix=".tmp",
            delete=False,
            newline="",
        ) as file:
            temp_path = Path(file.name)
            file.write(content)
            file.flush()
        temp_path.replace(destination)
    except Exception:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)
        raise

    return destination


def write_workbook_report(sheets: list[tuple[str, list[list[Any]]]], output_path: Path | str) -> Path:
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for title, rows in sheets:
        worksheet = workbook.create_sheet(title=_make_sheet_title(title, workbook.sheetnames))
        for row in rows:
            worksheet.append([_cell(value) for value in row])
        _format_worksheet(worksheet)

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(
            "wb",
            dir=str(destination.parent),
            prefix=f".{destination.name}.",
            suffix=".tmp",
            delete=False,
        ) as file:
            temp_path = Path(file.name)
            workbook.save(file.name)
        temp_path.replace(destination)
    except Exception:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)
        raise
    finally:
        workbook.close()

    return destination


def write_markdown_report(content: str, output_path: Path | str) -> Path:
    """Backward-compatible alias for older callers."""
    return write_csv_report(content, output_path)


def write_single_scan_report(
    result: dict[str, Any],
    source_path: Path | str,
    output_path: Path | str | None = None,
) -> Path:
    destination = _normalize_workbook_destination(output_path, source_path)
    return write_workbook_report(build_single_scan_workbook(result, source_path), destination)


def write_single_scan_report_from_file(
    source_path: Path | str,
    output_path: Path | str | None = None,
) -> Path:
    result = load_json_file(source_path)
    return write_single_scan_report(result, source_path, output_path)


def write_batch_summary_report(
    summary: dict[str, Any],
    source_path: Path | str,
    output_path: Path | str | None = None,
) -> Path:
    destination = _normalize_workbook_destination(output_path, source_path)
    return write_workbook_report(build_batch_summary_workbook(summary, source_path), destination)


def write_batch_item_reports(summary: dict[str, Any]) -> list[Path]:
    written: list[Path] = []
    for item in summary.get("items", []):
        saved_path = item.get("saved_path")
        if not saved_path:
            continue
        path = Path(saved_path)
        if not path.exists():
            continue
        written.append(write_single_scan_report_from_file(path))
    return written


def build_single_scan_report(result: dict[str, Any], source_path: Path | str | None = None) -> str:
    stats = result.get("statistics", {})
    wildcard = result.get("wildcard", {})
    tool_runs = result.get("tool_runs", {})
    port_scan = result.get("port_scan", {})
    web_fingerprint = result.get("web_fingerprint", {})
    directory_scan = result.get("directory_scan", {})

    rows: list[list[str]] = []
    order = 1

    def add(
        section: str,
        name: str,
        status: str = "",
        value: Any = "",
        detail_1: Any = "",
        detail_2: Any = "",
        detail_3: Any = "",
        note: Any = "",
    ):
        nonlocal order
        rows.append(
            [
                str(order),
                str(section),
                str(name),
                _cell(status),
                _cell(value),
                _cell(detail_1),
                _cell(detail_2),
                _cell(detail_3),
                _cell(note),
            ]
        )
        order += 1

    add("概览", "目标域名", value=result.get("target", "-"))
    add("概览", "结果文件", value=source_path or "-")
    add("概览", "扫描时间", value=result.get("scan_time", "-"))
    add("概览", "扫描耗时(秒)", value=result.get("duration_seconds", "-"))
    add("概览", "使用工具", value=", ".join(result.get("tools_used", [])) or "-")
    add("概览", "泛解析", status="是" if wildcard.get("detected") else "否", value=", ".join(wildcard.get("ips", [])))
    _add_blank_row(rows)

    add("统计", "原始发现数", value=stats.get("total_found", 0))
    add("统计", "有效子域名数", value=stats.get("valid_count", len(result.get("subdomains", []))))
    add("统计", "过滤数量", value=stats.get("filtered_count", 0))
    add("统计", "开放端口总数", value=port_scan.get("statistics", {}).get("total_open_ports", 0))
    add("统计", "Web 目标数", value=web_fingerprint.get("statistics", {}).get("web_target_count", 0))
    add("统计", "目录发现数", value=directory_scan.get("statistics", {}).get("interesting_path_count", 0))

    if tool_runs:
        _add_blank_row(rows)
        for tool_name, info in sorted(tool_runs.items()):
            add(
                "工具状态",
                tool_name,
                status=info.get("status", "-"),
                value=info.get("valid_count", 0),
                detail_1=f"原始 {info.get('raw_count', 0)}",
                detail_2=f"返回码 {info.get('return_code', '-')}",
                note=info.get("message", "-"),
            )

    subdomains = result.get("subdomains", [])
    if subdomains:
        _add_blank_row(rows)
        for item in subdomains:
            ips = ", ".join(item.get("ip", [])) or "-"
            add(
                "子域名",
                item.get("subdomain", "-"),
                value=len(item.get("ip", [])),
                detail_1=ips,
            )

    port_hosts = port_scan.get("hosts", {})
    if port_hosts:
        _add_blank_row(rows)
        for ip, ports in sorted(port_hosts.items()):
            add(
                "开放端口",
                ip,
                value=len(ports),
                detail_1=", ".join(str(port) for port in ports),
            )

    web_targets = web_fingerprint.get("targets", [])
    if web_targets:
        _add_blank_row(rows)
        for target in web_targets:
            nmap_info = target.get("nmap", {})
            add(
                "Web目标",
                target.get("url", "-"),
                status=target.get("fingerprint_status", ""),
                value=target.get("port", ""),
                detail_1=target.get("ip", ""),
                detail_2=nmap_info.get("title", ""),
                detail_3=nmap_info.get("server_header", ""),
            )

    findings = _flatten_directory_findings(directory_scan)
    if findings:
        _add_blank_row(rows)
        for item in findings:
            add(
                "目录发现",
                item["url"],
                status=item["status"],
                value=item["path"],
                detail_1=item["redirect"],
            )

    return _rows_to_csv(rows)


def build_batch_summary_report(summary: dict[str, Any], source_path: Path | str | None = None) -> str:
    stats = summary.get("statistics", {})
    items = summary.get("items", [])

    rows: list[list[str]] = []
    order = 1

    def add(
        section: str,
        name: str,
        status: str = "",
        value: Any = "",
        detail_1: Any = "",
        detail_2: Any = "",
        detail_3: Any = "",
        note: Any = "",
    ):
        nonlocal order
        rows.append(
            [
                str(order),
                str(section),
                str(name),
                _cell(status),
                _cell(value),
                _cell(detail_1),
                _cell(detail_2),
                _cell(detail_3),
                _cell(note),
            ]
        )
        order += 1

    add("概览", "汇总文件", value=source_path or "-")
    add("概览", "扫描时间", value=summary.get("scan_time", "-"))
    add("概览", "使用工具", value=", ".join(summary.get("tools_used", [])) or "-")
    _add_blank_row(rows)

    add("统计", "请求域名数", value=stats.get("requested_domains", stats.get("total_domains", 0)))
    add("统计", "成功数", value=stats.get("success_count", 0))
    add("统计", "无有效结果数", value=stats.get("no_valid_result_count", 0))
    add("统计", "错误数", value=stats.get("error_count", 0))
    add("统计", "开放端口总数", value=stats.get("total_open_ports", 0))
    add("统计", "Web 目标总数", value=stats.get("total_web_targets", 0))
    add("统计", "目录发现总数", value=stats.get("total_dirsearch_findings", 0))

    if items:
        _add_blank_row(rows)
        for item in items:
            add(
                "域名概览",
                item.get("domain", "-"),
                status=item.get("status", "-"),
                value=item.get("valid_count", 0),
                detail_1=f"开放端口 {item.get('open_port_count', 0)}",
                detail_2=f"Web目标 {item.get('web_target_count', 0)}",
                detail_3=f"目录发现 {item.get('dirsearch_finding_count', 0)}",
                note=item.get("message", "-"),
            )

    return _rows_to_csv(rows)


def build_single_scan_workbook(
    result: dict[str, Any],
    source_path: Path | str | None = None,
) -> list[tuple[str, list[list[Any]]]]:
    stats = result.get("statistics", {})
    wildcard = result.get("wildcard", {})
    tool_runs = result.get("tool_runs", {})
    subdomains = result.get("subdomains", [])
    port_scan = result.get("port_scan", {})
    web_fingerprint = result.get("web_fingerprint", {})
    directory_scan = result.get("directory_scan", {})

    sheets: list[tuple[str, list[list[Any]]]] = []
    sheets.append(
        (
            "概览",
            _sheet_rows(
                ["字段", "值"],
                [
                    ["目标域名", result.get("target", "-")],
                    ["结果文件", source_path or "-"],
                    ["扫描时间", result.get("scan_time", "-")],
                    ["扫描耗时(秒)", result.get("duration_seconds", "-")],
                    ["使用工具", ", ".join(result.get("tools_used", [])) or "-"],
                    ["泛解析", "是" if wildcard.get("detected") else "否"],
                    ["泛解析IP", ", ".join(wildcard.get("ips", [])) or "-"],
                ],
            ),
        )
    )
    sheets.append(
        (
            "统计",
            _sheet_rows(
                ["指标", "值"],
                [
                    ["原始发现数", stats.get("total_found", 0)],
                    ["有效子域名数", stats.get("valid_count", len(subdomains))],
                    ["过滤数量", stats.get("filtered_count", 0)],
                    ["开放端口总数", port_scan.get("statistics", {}).get("total_open_ports", 0)],
                    ["Web目标数", web_fingerprint.get("statistics", {}).get("web_target_count", 0)],
                    ["目录发现数", directory_scan.get("statistics", {}).get("interesting_path_count", 0)],
                ],
            ),
        )
    )
    sheets.append(
        (
            "工具状态",
            _sheet_rows(
                ["工具", "状态", "有效数", "原始数", "返回码", "消息"],
                [
                    [
                        tool_name,
                        info.get("status", "-"),
                        info.get("valid_count", 0),
                        info.get("raw_count", 0),
                        info.get("return_code", "-"),
                        info.get("message", "-"),
                    ]
                    for tool_name, info in sorted(tool_runs.items())
                ],
            ),
        )
    )
    sheets.append(
        (
            "子域名",
            _sheet_rows(
                ["子域名", "IP数量", "IP列表", "存活验证"],
                [
                    [
                        item.get("subdomain", "-"),
                        len(item.get("ip", [])),
                        ", ".join(item.get("ip", [])) or "-",
                        "是" if item.get("alive_verified") else "否",
                    ]
                    for item in subdomains
                ],
            ),
        )
    )

    port_hosts = port_scan.get("hosts", {})
    sheets.append(
        (
            "端口扫描",
            _sheet_rows(
                ["IP", "开放端口数", "端口列表"],
                [
                    [ip, len(ports), ", ".join(str(port) for port in ports)]
                    for ip, ports in sorted(port_hosts.items())
                ],
            ),
        )
    )

    web_targets = web_fingerprint.get("targets", [])
    sheets.append(
        (
            "Web指纹",
            _sheet_rows(
                ["URL", "子域名", "端口", "IP", "状态", "标题", "Server Header", "Service"],
                [
                    [
                        target.get("url", "-"),
                        target.get("subdomain", "-"),
                        target.get("port", ""),
                        target.get("ip", ""),
                        target.get("fingerprint_status", ""),
                        target.get("nmap", {}).get("title", ""),
                        target.get("nmap", {}).get("server_header", ""),
                        target.get("nmap", {}).get("service", ""),
                    ]
                    for target in web_targets
                ],
            ),
        )
    )

    sheets.append(
        (
            "目录扫描",
            _sheet_rows(
                ["URL", "路径", "状态码", "重定向", "子域名", "IP"],
                [
                    [
                        item["url"],
                        item["path"],
                        item["status"],
                        item["redirect"],
                        item["subdomain"],
                        item["ip"],
                    ]
                    for item in _flatten_directory_findings(directory_scan)
                ],
            ),
        )
    )
    return sheets


def build_batch_summary_workbook(
    summary: dict[str, Any],
    source_path: Path | str | None = None,
) -> list[tuple[str, list[list[Any]]]]:
    stats = summary.get("statistics", {})
    items = summary.get("items", [])

    return [
        (
            "概览",
            _sheet_rows(
                ["字段", "值"],
                [
                    ["汇总文件", source_path or "-"],
                    ["扫描时间", summary.get("scan_time", "-")],
                    ["使用工具", ", ".join(summary.get("tools_used", [])) or "-"],
                ],
            ),
        ),
        (
            "统计",
            _sheet_rows(
                ["指标", "值"],
                [
                    ["请求域名数", stats.get("requested_domains", stats.get("total_domains", 0))],
                    ["成功数", stats.get("success_count", 0)],
                    ["无有效结果数", stats.get("no_valid_result_count", 0)],
                    ["错误数", stats.get("error_count", 0)],
                    ["开放端口总数", stats.get("total_open_ports", 0)],
                    ["Web目标总数", stats.get("total_web_targets", 0)],
                    ["目录发现总数", stats.get("total_dirsearch_findings", 0)],
                ],
            ),
        ),
        (
            "域名概览",
            _sheet_rows(
                ["域名", "状态", "消息", "有效子域名数", "原始发现数", "泛解析", "结果文件"],
                [
                    [
                        item.get("domain", "-"),
                        item.get("status", "-"),
                        item.get("message", "-"),
                        item.get("valid_count", 0),
                        item.get("total_found", 0),
                        "是" if item.get("wildcard_detected") else "否",
                        item.get("saved_path", "-"),
                    ]
                    for item in items
                ],
            ),
        ),
        (
            "端口扫描",
            _sheet_rows(
                ["域名", "阶段状态", "开放端口数"],
                [
                    [
                        item.get("domain", "-"),
                        item.get("port_scan_status", "not_started"),
                        item.get("open_port_count", 0),
                    ]
                    for item in items
                ],
            ),
        ),
        (
            "Web指纹",
            _sheet_rows(
                ["域名", "阶段状态", "Web目标数"],
                [
                    [
                        item.get("domain", "-"),
                        item.get("web_fingerprint_status", "not_started"),
                        item.get("web_target_count", 0),
                    ]
                    for item in items
                ],
            ),
        ),
        (
            "目录扫描",
            _sheet_rows(
                ["域名", "阶段状态", "目录发现数"],
                [
                    [
                        item.get("domain", "-"),
                        item.get("directory_scan_status", "not_started"),
                        item.get("dirsearch_finding_count", 0),
                    ]
                    for item in items
                ],
            ),
        ),
    ]


def _rows_to_csv(rows: list[list[str]]) -> str:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)
    writer.writerow(CSV_HEADERS)
    writer.writerows(rows)
    return buffer.getvalue()


def _add_blank_row(rows: list[list[str]]):
    rows.append([""] * len(CSV_HEADERS))


def _sheet_rows(headers: list[str], rows: list[list[Any]]) -> list[list[Any]]:
    normalized_rows = rows or [["无数据", *([""] * (len(headers) - 1))]]
    return [headers, *normalized_rows]


def _normalize_workbook_destination(
    output_path: Path | str | None,
    source_path: Path | str,
) -> Path:
    if output_path is None:
        return default_report_path(source_path)

    destination = Path(output_path)
    if destination.suffix.lower() == ".xlsx":
        return destination
    return destination.with_suffix(".xlsx")


def _flatten_directory_findings(directory_scan: dict[str, Any]) -> list[dict[str, Any]]:
    findings: list[dict[str, Any]] = []
    for target in directory_scan.get("targets", []):
        url = str(target.get("url", "-"))
        for finding in target.get("findings", []):
            findings.append(
                {
                    "url": url,
                    "path": str(finding.get("path", "-")),
                    "status": int(finding.get("status", 0) or 0),
                    "redirect": str(finding.get("redirect", "") or "").strip(),
                    "subdomain": str(target.get("subdomain", "-") or "-"),
                    "ip": str(target.get("ip", "") or ""),
                }
            )
    return findings


def _format_worksheet(worksheet):
    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = WORKBOOK_HEADER_FILL
        cell.font = WORKBOOK_HEADER_FONT
        cell.alignment = WORKBOOK_CELL_ALIGNMENT

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WORKBOOK_CELL_ALIGNMENT

    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        width = min(max(len(value) for value in values) + 2, 60)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max(width, 12)


def _make_sheet_title(title: str, existing_titles: list[str]) -> str:
    sanitized = title.replace("/", "_").replace("\\", "_").replace(":", "：").strip() or "Sheet"
    sanitized = sanitized[:31]
    if sanitized not in existing_titles:
        return sanitized

    base = sanitized[:28] or "Sheet"
    index = 2
    while True:
        candidate = f"{base}-{index}"[:31]
        if candidate not in existing_titles:
            return candidate
        index += 1


def _cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip()
